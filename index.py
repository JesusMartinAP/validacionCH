import concurrent.futures
import tkinter as tk
from tkinter import messagebox
from datetime import datetime
from playwright.sync_api import sync_playwright
from openpyxl import Workbook
import webbrowser
from tkinter import ttk
import threading
import logging
import subprocess

# Configuración de logs
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')

# Variables globales
proceso_en_ejecucion = False
estado_codigos = []
total_codigos = 0
codigos_procesados = 0
lock = threading.Lock()  # Para controlar el acceso concurrente a las variables globales

def ensure_playwright_browsers_installed():
    try:
        with sync_playwright() as p:
            # Intenta abrir el navegador para verificar si está instalado
            p.chromium.launch(headless=True).close()
    except Exception as e:
        # Si falla, intenta instalar los navegadores
        if "Executable doesn't exist" in str(e):
            print("Instalando navegadores de Playwright...")
            subprocess.run(["playwright", "install"], check=True)

ensure_playwright_browsers_installed()

# Función para obtener el estado de un producto, su precio y la cantidad de imágenes usando Playwright
def obtener_estado_y_precio(codigo_padre):
    url_base = f'https://www.marathon.cl/{codigo_padre}.html'
    with sync_playwright() as p:
        browser = p.chromium.launch(headless=True)
        page = browser.new_page()
        try:
            page.goto(url_base)
            page.wait_for_selector("body", timeout=5000)  # Esperar a que el cuerpo de la página se cargue

            # Verificar si la página redirige al inicio (indicador: URL de redirección)
            if page.url == "https://www.marathon.cl/home/":
                browser.close()
                logging.warning(f'Web no encontrada para código: {codigo_padre}')
                return codigo_padre, "Web no encontrada", "Precio no disponible", "Cantidad de imágenes no disponible"

            # Buscar los botones de talla y determinar si están seleccionados
            botones_talla = page.query_selector_all("button.size-attribute.swatchable.selectable.swatch-square")
            estado = "Agotado"
            for boton in botones_talla:
                if "selected-assistive-text" in boton.inner_html():
                    estado = "Disponible" if not boton.is_disabled() else "Agotado"
                    break

            # Extraer el precio
            try:
                precio_element = page.query_selector('span.sales > span.value')
                precio = precio_element.inner_text().strip() if precio_element else "Precio no disponible"
            except:
                precio = "Precio no disponible"
            
            # Contar la cantidad de imágenes
            try:
                imagenes = page.query_selector_all('img.galley_img')
                cantidad_imagenes = len(imagenes)
            except:
                cantidad_imagenes = "Cantidad de imágenes no disponible"
            
            browser.close()
            logging.info(f'Código {codigo_padre} procesado correctamente.')
            return codigo_padre, estado, precio, cantidad_imagenes
        except Exception as e:
            browser.close()
            logging.error(f'Error procesando código {codigo_padre}: {e}')
            return codigo_padre, f"Error: {e}", "Precio no disponible", "Cantidad de imágenes no disponible"

# Función para procesar los códigos y guardar en Excel
def procesar_codigos(codigos):
    global proceso_en_ejecucion, estado_codigos, total_codigos, codigos_procesados
    total_codigos = len(codigos)
    start_time = datetime.now()

    def obtener_estado_concurrente(codigo_padre):
        return obtener_estado_y_precio(codigo_padre)
    
    with concurrent.futures.ThreadPoolExecutor(max_workers=50) as executor:  # Ajustar el número de trabajadores según sea necesario
        futures = {executor.submit(obtener_estado_concurrente, codigo): codigo for codigo in codigos}
        for future in concurrent.futures.as_completed(futures):
            if not proceso_en_ejecucion:
                break
            codigo_padre = futures[future]  # Asignar código_padre aquí para asegurarse de que siempre tenga un valor
            try:
                codigo_padre, estado, precio, cantidad_imagenes = future.result()
            except Exception as exc:
                estado = f"Error: {exc}"
                precio = "Precio no disponible"
                cantidad_imagenes = "Cantidad de imágenes no disponible"
            with lock:
                estado_codigos.append((codigo_padre, estado, precio, cantidad_imagenes))
                codigos_procesados += 1

            # Actualizar información en la interfaz
            elapsed_time = datetime.now() - start_time
            tiempo_transcurrido = str(elapsed_time).split('.')[0]  # Formato HH:MM:SS
            info_estado.set(f"Procesando código {codigos_procesados}/{total_codigos} - Tiempo transcurrido: {tiempo_transcurrido}")
            barra_progreso['value'] = (codigos_procesados / total_codigos) * 100
            root.update_idletasks()

    if proceso_en_ejecucion:
        guardar_resultados()

# Función para pausar el proceso
def pausar_proceso():
    global proceso_en_ejecucion
    proceso_en_ejecucion = False
    messagebox.showinfo("Proceso pausado", "Se ha pausado el proceso. Puede continuar luego.")

# Función para detener el proceso
def detener_proceso():
    global proceso_en_ejecucion
    proceso_en_ejecucion = False
    guardar_resultados()
    messagebox.showinfo("Proceso detenido", "Se ha detenido el proceso.")
    root.quit()

# Función para guardar los resultados en Excel
def guardar_resultados():
    global proceso_en_ejecucion
    # Guardar en Excel
    wb = Workbook()
    ws = wb.active
    ws.title = "Control Stock Web"
    ws['A1'] = "CODIGO"
    ws['B1'] = "STATUS WEB"
    ws['C1'] = "PRECIO"
    ws['D1'] = "Cant. Img"
    for i, (codigo, estado, precio, cantidad_imagenes) in enumerate(estado_codigos, start=2):
        ws[f'A{i}'] = codigo
        ws[f'B{i}'] = estado
        ws[f'C{i}'] = precio
        ws[f'D{i}'] = cantidad_imagenes

    # Guardar archivo
    fecha_actual = datetime.now().strftime("%Y-%m-%d")
    nombre_archivo = f"Control Stock Web {fecha_actual}.xlsx"
    wb.save(nombre_archivo)
    messagebox.showinfo("Proceso completado", f"Se ha guardado el archivo '{nombre_archivo}' con los estados, precios y cantidad de imágenes de los productos.")
    # Abrir archivo Excel
    webbrowser.open(nombre_archivo)

# Función para iniciar el procesamiento en un hilo separado
def iniciar_procesamiento():
    global proceso_en_ejecucion, codigos_procesados, estado_codigos
    if not proceso_en_ejecucion:
        proceso_en_ejecucion = True
        codigos = entry_codigos.get("1.0", "end").strip().split()  # Añadir .strip() para eliminar espacios en blanco al final
        codigos_procesados = 0
        estado_codigos = []
        threading.Thread(target=procesar_codigos, args=(codigos,)).start()

# Interfaz gráfica
root = tk.Tk()
root.title("Verificar Stock Web")

# Entrada de códigos
tk.Label(root, text="Ingrese los códigos separados por espacio:").pack()
entry_codigos = tk.Text(root, height=10, width=50)
entry_codigos.pack()

# Botón de iniciar procesamiento
tk.Button(root, text="Iniciar Procesamiento", command=iniciar_procesamiento).pack()

# Botón de detener
btn_detener = tk.Button(root, text="Detener", command=detener_proceso)
btn_detener.pack()

# Información de estado y progreso
info_estado = tk.StringVar()
tk.Label(root, textvariable=info_estado).pack()

barra_progreso = ttk.Progressbar(root, length=200, mode='determinate')
barra_progreso.pack()

root.mainloop()
